import os
import datetime
import uuid
from openpyxl import Workbook, load_workbook
from io import BytesIO

def generate_user_id(user_name):
    """
    Generate a unique user ID based on the user's name.
    
    Args:
        user_name (str): The full name of the user.
    
    Returns:
        str: A unique user ID in the format 'cleaned_name_randomsuffix'.
    """
    import uuid
    clean_name = "".join(c for c in user_name.lower() if c.isalnum())
    return f"{clean_name}_{uuid.uuid4().hex[:8]}"


def save_user_data(user_id, voice_id, name=None, email=None):
    path = "data/User_Data.xlsx"
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.append(["User_ID", "Voice_ID", "Name", "Email", "Timestamp"])
        wb.save(path)

    wb = load_workbook(path)
    ws = wb.active
    ws.append([user_id, voice_id, name or "", email or "", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    wb.save(path)

def load_existing_users():
    path = "data/User_Data.xlsx"
    users = {}
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            users[row[0]] = row[1]  # user_id: voice_id
    return users

def save_text_template():
    wb = Workbook()
    ws = wb.active
    ws.append(["Text", "File_name"])
    ws.append(["Hello, how are you?", "greeting1"])
    ws.append(["Welcome to our app!", "welcome1"])
    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)
    virtual_workbook.seek(0)
    return virtual_workbook

def load_text_inputs(file=None, custom_text=None):
    texts = {}
    if file:
        wb = load_workbook(file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                texts[row[1]] = row[0]  # file_name: text
    if custom_text:
        texts["custom"] = custom_text
    return texts
